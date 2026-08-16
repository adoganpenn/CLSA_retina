"""Reusable Databricks helpers for the CLSA retinal-aging extraction pipeline.

The module is intentionally path-driven: it inventories a Unity Catalog Volume,
harmonizes the Tracking and Comprehensive Follow-up 2 cohorts, derives the
pre-specified retinal-aging variables, and validates genetics-file readiness.
It never stores archive passwords in source code or output tables.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
import csv
import hashlib
import json
import os
from pathlib import Path
import re
from typing import Any, Iterable, Mapping, Sequence
import zipfile


STRUCTURAL_MISSING_CODES = (
    "-77771",
    "-77772",
    "-88880",
    "-88888",
    "-99991",
    "-99993",
    "-99999",
)

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp", ".dcm"}


@dataclass(frozen=True)
class VariableSpec:
    standard_name: str
    measure: str
    tracking_candidates: tuple[str, ...]
    comprehensive_candidates: tuple[str, ...]
    baseline_candidates: tuple[str, ...]
    value_kind: str
    role: str
    status: str
    notes: str


def load_json(path: str | os.PathLike[str]) -> dict[str, Any]:
    with open(path, encoding="utf-8") as stream:
        return json.load(stream)


def _split_candidates(value: str | None) -> tuple[str, ...]:
    return tuple(item.strip() for item in (value or "").split("|") if item.strip())


def load_variable_manifest(path: str | os.PathLike[str]) -> list[VariableSpec]:
    with open(path, newline="", encoding="utf-8") as stream:
        rows = csv.DictReader(stream)
        return [
            VariableSpec(
                standard_name=row["standard_name"].strip(),
                measure=row["measure"].strip(),
                tracking_candidates=_split_candidates(row["tracking_candidates"]),
                comprehensive_candidates=_split_candidates(
                    row["comprehensive_candidates"]
                ),
                baseline_candidates=_split_candidates(row["baseline_candidates"]),
                value_kind=row["value_kind"].strip(),
                role=row["role"].strip(),
                status=row["status"].strip(),
                notes=row["notes"].strip(),
            )
            for row in rows
        ]


def build_file_inventory(spark: Any, volume_root: str) -> Any:
    """Return a Spark DataFrame containing metadata only (never file contents)."""
    from pyspark.sql import functions as F

    root = volume_root.rstrip("/")
    root_pattern = rf"^(?:dbfs:)?{re.escape(root)}/?"
    files = (
        spark.read.format("binaryFile")
        .option("recursiveFileLookup", "true")
        .load(root)
        .select(
            "path",
            F.col("length").cast("long").alias("bytes"),
            F.col("modificationTime").alias("modified_utc"),
        )
    )
    name = F.element_at(F.split(F.col("path"), "/"), -1)
    lower_name = F.lower(name)
    extension = (
        F.when(lower_name.endswith(".bgen.bgi"), F.lit(".bgen.bgi"))
        .when(lower_name.endswith(".csv.gz"), F.lit(".csv.gz"))
        .when(lower_name.endswith(".txt.gz"), F.lit(".txt.gz"))
        .otherwise(
            F.concat(
                F.lit("."),
                F.regexp_extract(lower_name, r"\.([^.]+)$", 1),
            )
        )
    )
    return files.select(
        F.regexp_replace(F.col("path"), r"^dbfs:", "").alias("path"),
        F.regexp_replace(F.col("path"), root_pattern, "").alias("relative_path"),
        name.alias("name"),
        extension.alias("extension"),
        "bytes",
        "modified_utc",
    )


def discover_paths(
    inventory_df: Any, regex: str, *, path_column: str = "relative_path"
) -> list[str]:
    from pyspark.sql import functions as F

    return [
        row["path"]
        for row in inventory_df.filter(F.col(path_column).rlike(regex))
        .select("path")
        .orderBy("path")
        .collect()
    ]


def _safe_archive_destination(output_root: Path, member_name: str) -> Path:
    destination = (output_root / member_name).resolve()
    resolved_root = output_root.resolve()
    if resolved_root != destination and resolved_root not in destination.parents:
        raise ValueError(f"Unsafe archive member path: {member_name}")
    return destination


def extract_zip_archive(
    archive_path: str,
    output_root: str,
    password: str,
    *,
    overwrite: bool = False,
) -> list[dict[str, Any]]:
    """Safely extract a ZipCrypto archive and return an extraction manifest.

    Python's standard library does not decrypt WinZip AES archives. If the CLSA
    archive uses AES, install and invoke a reviewed 7-Zip binary on the cluster
    instead of placing the password on a shell command line.
    """
    destination_root = Path(output_root)
    destination_root.mkdir(parents=True, exist_ok=True)
    manifest: list[dict[str, Any]] = []
    password_bytes = password.encode("utf-8")

    with zipfile.ZipFile(archive_path) as archive:
        for info in archive.infolist():
            destination = _safe_archive_destination(destination_root, info.filename)
            if info.is_dir():
                destination.mkdir(parents=True, exist_ok=True)
                continue
            destination.parent.mkdir(parents=True, exist_ok=True)
            status = "extracted"
            if (
                destination.exists()
                and destination.stat().st_size == info.file_size
                and not overwrite
            ):
                status = "already_present"
            else:
                try:
                    with archive.open(info, pwd=password_bytes) as source:
                        with open(destination, "wb") as target:
                            while True:
                                chunk = source.read(8 * 1024 * 1024)
                                if not chunk:
                                    break
                                target.write(chunk)
                except NotImplementedError as exc:
                    raise RuntimeError(
                        "Archive encryption is unsupported by Python zipfile. "
                        "Use an approved AES-capable extractor with secret-safe "
                        "password handling."
                    ) from exc
            manifest.append(
                {
                    "archive_path": archive_path,
                    "member_path": info.filename,
                    "output_path": str(destination),
                    "uncompressed_bytes": int(info.file_size),
                    "crc32": f"{info.CRC:08x}",
                    "status": status,
                }
            )
    return manifest


def extract_fundus_archives_from_secret(
    dbutils: Any,
    archive_paths: Sequence[str],
    output_root: str,
    *,
    secret_scope: str,
    secret_key: str,
    overwrite: bool = False,
) -> list[dict[str, Any]]:
    password = dbutils.secrets.get(scope=secret_scope, key=secret_key)
    try:
        all_rows: list[dict[str, Any]] = []
        for archive_path in archive_paths:
            if not archive_path.lower().endswith(".zip"):
                raise ValueError(
                    f"Only .zip is supported by the Python extractor: {archive_path}"
                )
            all_rows.extend(
                extract_zip_archive(
                    archive_path, output_root, password, overwrite=overwrite
                )
            )
        return all_rows
    finally:
        password = ""


def _first_named_group(regex: re.Pattern[str], value: str, group: str) -> str | None:
    match = regex.search(value)
    return match.groupdict().get(group) if match else None


def build_image_manifest(
    spark: Any,
    image_root: str,
    *,
    participant_id_regex: str,
    eye_regex: str,
    probe_dimensions: bool = False,
    compute_sha256: bool = False,
) -> Any:
    root = Path(image_root)
    id_pattern = re.compile(participant_id_regex)
    eye_pattern = re.compile(eye_regex)
    rows: list[dict[str, Any]] = []

    for directory, _, filenames in os.walk(root):
        for filename in filenames:
            path = Path(directory) / filename
            if path.suffix.lower() not in IMAGE_EXTENSIONS:
                continue
            relative_path = str(path.relative_to(root))
            stat = path.stat()
            width = None
            height = None
            if probe_dimensions and path.suffix.lower() != ".dcm":
                try:
                    from PIL import Image

                    with Image.open(path) as image:
                        width, height = image.size
                except Exception:
                    width, height = None, None
            digest = None
            if compute_sha256:
                hasher = hashlib.sha256()
                with open(path, "rb") as stream:
                    for chunk in iter(lambda: stream.read(8 * 1024 * 1024), b""):
                        hasher.update(chunk)
                digest = hasher.hexdigest()
            participant_id = _first_named_group(
                id_pattern, relative_path, "participant_id"
            )
            eye = _first_named_group(eye_pattern, relative_path, "eye")
            if eye:
                eye = {
                    "R": "R",
                    "RIGHT": "R",
                    "OD": "R",
                    "L": "L",
                    "LEFT": "L",
                    "OS": "L",
                    "OU": "BOTH",
                }.get(eye.upper(), eye.upper())
            rows.append(
                {
                    "image_path": str(path),
                    "relative_path": relative_path,
                    "filename": filename,
                    "extension": path.suffix.lower(),
                    "bytes": int(stat.st_size),
                    "modified_utc": datetime.fromtimestamp(
                        stat.st_mtime, tz=timezone.utc
                    ),
                    "participant_id_parsed": participant_id,
                    "eye_parsed": eye,
                    "width_px": width,
                    "height_px": height,
                    "sha256": digest,
                    "parse_ok": participant_id is not None,
                }
            )

    schema = (
        "image_path string, relative_path string, filename string, extension string, "
        "bytes long, modified_utc timestamp, participant_id_parsed string, "
        "eye_parsed string, width_px int, height_px int, sha256 string, parse_ok boolean"
    )
    return spark.createDataFrame(rows, schema=schema)


def load_dictionary_sheets(
    spark: Any, workbook_path: str
) -> tuple[Any, Any]:
    """Load the Variables and Categories sheets into two Spark DataFrames."""
    import openpyxl
    from pyspark.sql.types import StringType, StructField, StructType

    workbook = openpyxl.load_workbook(
        workbook_path, read_only=True, data_only=True
    )

    def sheet_records(sheet_name: str) -> list[dict[str, str | None]]:
        sheet = workbook[sheet_name]
        headers = [
            str(cell.value) if cell.value is not None else f"column_{index}"
            for index, cell in enumerate(
                next(sheet.iter_rows(min_row=1, max_row=1)), start=1
            )
        ]
        return [
            {
                header: None if value is None else str(value)
                for header, value in zip(headers, row)
            }
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if any(value is not None for value in row)
        ]

    def sheet_data_frame(sheet_name: str) -> Any:
        records = sheet_records(sheet_name)
        if not records:
            raise ValueError(f"Dictionary sheet is empty: {sheet_name}")
        schema = StructType(
            [
                StructField(header, StringType(), nullable=True)
                for header in records[0]
            ]
        )
        return spark.createDataFrame(records, schema=schema)

    try:
        variables = sheet_data_frame("Variables")
        categories = sheet_data_frame("Categories")
    finally:
        workbook.close()
    return variables, categories


def read_tabular_auto(
    spark: Any,
    path: str,
    *,
    csv_options: Mapping[str, str] | None = None,
) -> Any:
    lower = path.lower().rstrip("/")
    if lower.endswith(".parquet"):
        return spark.read.parquet(path)
    if lower.endswith(".delta") or Path(path).is_dir():
        try:
            return spark.read.format("delta").load(path)
        except Exception:
            if Path(path).is_dir():
                return spark.read.parquet(path)
            raise
    if lower.endswith(".csv") or lower.endswith(".csv.gz"):
        options = {
            "header": "true",
            "inferSchema": "false",
            "multiLine": "true",
            "escape": '"',
        }
        options.update(csv_options or {})
        return spark.read.options(**options).csv(path)
    raise ValueError(
        f"Unsupported questionnaire format for {path}. Convert SAS/SPSS/Excel "
        "data extracts to CSV, Parquet, or Delta before harmonization."
    )


def read_whitespace_table(spark: Any, path: str) -> Any:
    """Read a modest whitespace-delimited genetics metadata file.

    CLSA sample/marker QC files are small enough for a pandas parse on the
    driver. Genotype BED/BGEN files must not be read through this helper.
    """
    import pandas as pd

    with open(path, encoding="utf-8", errors="replace") as stream:
        first_line = stream.readline().strip()
    first_token = first_line.split()[0] if first_line else ""
    has_header = bool(re.search(r"[A-Za-z_]", first_token))
    frame = pd.read_csv(
        path,
        sep=r"\s+",
        dtype=str,
        header=0 if has_header else None,
    )
    if not has_header:
        frame.columns = [
            f"column_{index}" for index in range(1, len(frame.columns) + 1)
        ]
    return spark.createDataFrame(frame)


def normalize_column_names(df: Any) -> Any:
    """Normalize punctuation-heavy source headers without changing values."""
    normalized = []
    seen: set[str] = set()
    for original in df.columns:
        candidate = re.sub(r"[^A-Za-z0-9]+", "_", original).strip("_")
        candidate = candidate or "column"
        base = candidate
        index = 2
        while candidate.casefold() in seen:
            candidate = f"{base}_{index}"
            index += 1
        seen.add(candidate.casefold())
        normalized.append(candidate)
    return df.toDF(*normalized)


def resolve_id_column(df: Any, candidates: Sequence[str]) -> str:
    exact = set(df.columns)
    casefolded = {column.casefold(): column for column in df.columns}
    for candidate in candidates:
        if candidate in exact:
            return candidate
        if candidate.casefold() in casefolded:
            return casefolded[candidate.casefold()]
    raise ValueError(
        "No participant identifier column was found. Available columns start "
        f"with: {df.columns[:20]}. Supply the authorized project-specific ID "
        "column explicitly; do not assume ADM_GWAS_COM is present."
    )


def _candidate_list(spec: VariableSpec, cohort: str) -> tuple[str, ...]:
    if cohort == "tracking":
        return spec.tracking_candidates
    if cohort == "comprehensive":
        return spec.comprehensive_candidates
    if cohort == "baseline":
        return spec.baseline_candidates
    raise ValueError(f"Unsupported cohort: {cohort}")


def harmonize_cohort(
    df: Any,
    *,
    cohort: str,
    manifest: Sequence[VariableSpec],
    id_candidates: Sequence[str],
) -> tuple[Any, list[dict[str, str | None]]]:
    """Project one source cohort to the manifest's common, string-typed schema."""
    from pyspark.sql import functions as F

    id_column = resolve_id_column(df, id_candidates)
    selections = [
        F.col(id_column).cast("string").alias("participant_id"),
        F.lit(cohort).alias("cohort"),
    ]
    audit: list[dict[str, str | None]] = []
    available = set(df.columns)
    casefolded = {column.casefold(): column for column in df.columns}

    for spec in manifest:
        selected = None
        for candidate in _candidate_list(spec, cohort):
            if candidate in available:
                selected = candidate
                break
            selected = casefolded.get(candidate.casefold())
            if selected:
                break
        if selected:
            selections.append(
                F.col(selected).cast("string").alias(spec.standard_name)
            )
        else:
            selections.append(F.lit(None).cast("string").alias(spec.standard_name))
        audit.append(
            {
                "cohort": cohort,
                "standard_name": spec.standard_name,
                "selected_source_column": selected,
                "status": "found" if selected else "missing",
            }
        )
    return df.select(*selections), audit


def _clean_code(
    column: Any, *, extra_missing_codes: Sequence[str] = ()
) -> Any:
    from pyspark.sql import functions as F

    normalized = F.regexp_replace(F.trim(column.cast("string")), r"\.0$", "")
    missing_codes = (*STRUCTURAL_MISSING_CODES, *extra_missing_codes)
    return F.when(normalized.isin(*missing_codes), F.lit(None)).otherwise(normalized)


def _clean_numeric(
    column: Any, *, extra_missing_codes: Sequence[str] = ()
) -> Any:
    return _clean_code(
        column, extra_missing_codes=extra_missing_codes
    ).cast("double")


def _binary_yes_no(column: Any) -> Any:
    from pyspark.sql import functions as F

    value = F.upper(_clean_code(column))
    return (
        F.when(value.isin("1", "11", "Y", "YES", "TRUE"), F.lit(1))
        .when(value.isin("0", "2", "N", "NO", "FALSE"), F.lit(0))
        .otherwise(F.lit(None).cast("int"))
    )


def _any_positive(columns: Sequence[Any]) -> Any:
    from pyspark.sql import functions as F

    any_observed = None
    any_yes = None
    for column in columns:
        observed = column.isNotNull()
        positive = column == 1
        any_observed = observed if any_observed is None else (any_observed | observed)
        any_yes = positive if any_yes is None else (any_yes | positive)
    return (
        F.when(any_yes, F.lit(1))
        .when(any_observed, F.lit(0))
        .otherwise(F.lit(None).cast("int"))
    )


def derive_retinal_metrics(
    harmonized_df: Any,
    *,
    visual_acuity_scale_confirmed_logmar: bool = False,
    require_both_eyes_for_better_eye: bool = True,
) -> Any:
    from pyspark.sql import functions as F

    df = harmonized_df
    categorical = [
        "self_reported_vision",
        "sex_at_birth",
        "ethnicity_spirometry",
        "education_level_sap",
        "marital_status",
        "household_income_band",
        "smoking_status",
        "social_outside_household",
        "social_religious",
        "social_education_culture",
        "social_club",
        "social_association",
        "social_other",
        "adl_class",
        "self_rated_healthy_aging",
        "sampling_strata",
    ]
    categorical_missing_8_9 = {
        "self_reported_vision",
        "education_level_sap",
        "marital_status",
        "household_income_band",
        "social_outside_household",
        "social_religious",
        "social_education_culture",
        "social_club",
        "social_association",
        "social_other",
        "self_rated_healthy_aging",
    }
    numeric = [
        "visual_acuity_left",
        "visual_acuity_right",
        "visual_acuity_both",
        "cesd10_score",
        "age_years",
        "analytic_weight",
        "frailty",
        "epigenetic_age",
        "epigenetic_age_acceleration_difference",
        "epigenetic_age_acceleration_residual",
        "epigenetic_ieaa",
        "epigenetic_eeaa",
        "epigenetic_hannum_age",
    ]
    binary = [
        "diabetes",
        "hypertension",
        "heart_disease",
        "stroke",
        "oa_hand",
        "oa_hip",
        "oa_knee",
        "osteoporosis",
        "asthma",
        "copd",
        "cancer",
        "low_back_pain",
        "hearing_noise",
        "hearing_aid",
    ]
    for name in categorical:
        extra_missing_codes = (
            ("8", "9") if name in categorical_missing_8_9 else ()
        )
        df = df.withColumn(
            name,
            _clean_code(
                F.col(name), extra_missing_codes=extra_missing_codes
            ),
        )
    for name in numeric:
        df = df.withColumn(name, _clean_numeric(F.col(name)))
    for name in binary:
        df = df.withColumn(name, _binary_yes_no(F.col(name)))

    vision_code = _clean_numeric(F.col("self_reported_vision"))
    df = df.withColumn(
        "visual_impairment_self_report",
        F.when(vision_code.isin(4.0, 5.0), F.lit(1))
        .when(vision_code.isin(1.0, 2.0, 3.0), F.lit(0))
        .otherwise(F.lit(None).cast("int")),
    )
    df = df.withColumn(
        "depression_cesd10",
        F.when(F.col("cesd10_score").isNull(), F.lit(None).cast("int"))
        .when(F.col("cesd10_score") >= 10, F.lit(1))
        .otherwise(F.lit(0)),
    )
    df = df.withColumn(
        "married_or_partnered",
        F.when(F.col("marital_status") == "2", F.lit(1))
        .when(F.col("marital_status").isin("1", "3", "4", "5"), F.lit(0))
        .otherwise(F.lit(None).cast("int")),
    )

    left = F.col("visual_acuity_left")
    right = F.col("visual_acuity_right")
    if require_both_eyes_for_better_eye:
        better_eye = F.when(left.isNotNull() & right.isNotNull(), F.least(left, right))
    else:
        better_eye = F.least(left, right)
    df = df.withColumn("visual_acuity_better_eye", better_eye)
    if visual_acuity_scale_confirmed_logmar:
        df = df.withColumn(
            "visual_impairment_acuity",
            F.when(
                F.col("visual_acuity_better_eye").isNull(),
                F.lit(None).cast("int"),
            )
            .when(F.col("visual_acuity_better_eye") > 0.3, F.lit(1))
            .otherwise(F.lit(0)),
        )
    else:
        df = df.withColumn(
            "visual_impairment_acuity", F.lit(None).cast("int")
        )

    df = df.withColumn(
        "arthritis_any",
        _any_positive([F.col("oa_hand"), F.col("oa_hip"), F.col("oa_knee")]),
    )
    df = df.withColumn(
        "asthma_or_copd", _any_positive([F.col("asthma"), F.col("copd")])
    )

    social_names = [
        "social_outside_household",
        "social_religious",
        "social_education_culture",
        "social_club",
        "social_association",
        "social_other",
    ]
    social_flags = []
    for name in social_names:
        code = _clean_numeric(F.col(name))
        flag_name = f"{name}_at_least_weekly"
        df = df.withColumn(
            flag_name,
            F.when(code.isin(1.0, 2.0), F.lit(1))
            .when(code.isin(3.0, 4.0, 5.0), F.lit(0))
            .otherwise(F.lit(None).cast("int")),
        )
        social_flags.append(F.col(flag_name))
    df = df.withColumn("social_any_at_least_weekly", _any_positive(social_flags))

    condition_names = [
        "diabetes",
        "hypertension",
        "heart_disease",
        "stroke",
        "arthritis_any",
        "osteoporosis",
        "asthma_or_copd",
        "cancer",
        "low_back_pain",
    ]
    observed = None
    count = None
    for name in condition_names:
        column = F.col(name)
        observed = (
            column.isNotNull() if observed is None else (observed | column.isNotNull())
        )
        count = (
            F.coalesce(column, F.lit(0))
            if count is None
            else count + F.coalesce(column, F.lit(0))
        )
    df = df.withColumn(
        "multimorbidity_selected_count",
        F.when(observed, count).otherwise(F.lit(None).cast("int")),
    )
    return df


def dictionary_missing_code_map(categories_df: Any) -> Any:
    """Return variable-specific missing codes as a compact Spark DataFrame."""
    from pyspark.sql import functions as F

    return (
        categories_df.filter(F.col("missing") == "1")
        .select(
            F.col("table"),
            F.col("variable"),
            F.col("name").alias("missing_code"),
            F.col("label:en").alias("missing_label"),
        )
        .dropDuplicates()
    )


def validate_genetics_inventory(
    inventory_df: Any, expected_chromosomes: Iterable[int] = range(1, 24)
) -> dict[str, Any]:
    names = {row["name"] for row in inventory_df.select("name").collect()}
    expected_chromosomes = list(expected_chromosomes)
    direct = {
        extension: f"clsa_gen_v3.{extension}" in names
        for extension in ("bed", "bim", "fam")
    }
    missing_bgen = [
        chromosome
        for chromosome in expected_chromosomes
        if f"clsa_imp_{chromosome}_v3.bgen" not in names
    ]
    missing_bgi = [
        chromosome
        for chromosome in expected_chromosomes
        if f"clsa_imp_{chromosome}_v3.bgen.bgi" not in names
    ]
    required_metadata = [
        "clsa_imp_v3.sample",
        "clsa_sqc_v3.txt",
        "clsa_mqc_v3.txt",
        "clsa_rel_v3.txt",
        "clsa_hla_v3.csv",
        "clsa_v3.md5",
    ]
    missing_metadata = [name for name in required_metadata if name not in names]
    return {
        "direct_genotypes": direct,
        "direct_ready": all(direct.values()),
        "missing_bgen_chromosomes": missing_bgen,
        "missing_bgi_chromosomes": missing_bgi,
        "imputed_ready": not missing_bgen and not missing_bgi,
        "missing_metadata": missing_metadata,
    }


def add_join_qc(
    df: Any,
    *,
    image_path_column: str = "image_path",
    genetics_id_column: str = "ADM_GWAS_COM",
) -> Any:
    from pyspark.sql import functions as F

    return (
        df.withColumn(
            "has_fundus_image", F.col(image_path_column).isNotNull().cast("int")
        )
        .withColumn(
            "has_genetics_link", F.col(genetics_id_column).isNotNull().cast("int")
        )
        .withColumn(
            "analysis_complete_case",
            (
                F.col("has_fundus_image").eqNullSafe(1)
                & F.col("has_genetics_link").eqNullSafe(1)
                & F.col("age_years").isNotNull()
                & F.col("sex_at_birth").isNotNull()
            ).cast("int"),
        )
    )


def write_delta(
    df: Any,
    path: str,
    *,
    mode: str = "overwrite",
    partition_by: Sequence[str] = (),
) -> None:
    writer = df.write.format("delta").mode(mode).option(
        "overwriteSchema", "true"
    )
    if partition_by:
        writer = writer.partitionBy(*partition_by)
    writer.save(path)
